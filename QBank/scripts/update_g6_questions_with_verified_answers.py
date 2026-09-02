import sys
import os
import re
import math
import json
import random
import sqlite3
import openpyxl

# Force UTF-8 output encoding
sys.stdout.reconfigure(encoding='utf-8')

print("🚀 Starting Grade 6 Question Bank Update using verified answers from Excel...")

excel_path = os.path.join('QBank', 'resources', 'Grade_6_Math_50_Unique_Questions_Per_With_Answers.xlsx')
if not os.path.exists(excel_path):
    print(f"Error: {excel_path} not found!")
    sys.exit(1)

wb = openpyxl.load_workbook(excel_path)
ws = wb['All 750 Questions']

items = []
for r in range(5, ws.max_row + 1):
    iid = str(ws.cell(r, 1).value or '').strip()
    if iid.startswith('T'):
        tcode = iid.split('-')[0]
        topic_name = str(ws.cell(r, 2).value or '').strip()
        domain = str(ws.cell(r, 3).value or '').strip()
        comp = str(ws.cell(r, 4).value or '').strip()
        stmt = str(ws.cell(r, 5).value or '').strip()
        ans = str(ws.cell(r, 6).value or '').strip()
        expl = str(ws.cell(r, 7).value or '').strip()
        
        items.append({
            'item_id': iid,
            'tcode': tcode,
            'topic_name': topic_name,
            'domain': domain,
            'comp': comp,
            'stmt': stmt,
            'ans': ans,
            'expl': expl
        })

print(f"Loaded {len(items)} seed questions with verified answers from {excel_path}")

topic_map = {
    'T01': 89,   # Tessellation of shapes
    'T02': 90,   # Translation, reflection and rotation with shapes
    'T03': 91,   # Units of volume and capacity
    'T04': 92,   # Volume of cubes and rectangular prisms
    'T05': 93,   # Perimeter and area of triangles, parallelograms, trapezoids, and composite figures
    'T06': 94,   # Parts of a circle, including circumference
    'T07': 95,   # Area of a circle
    'T08': 96,   # Composite figures (triangles, squares, rectangles, circles, semicircles)
    'T09': 97,   # The four operations with decimals
    'T10': 98,   # The four operations with fractions, whole numbers, and mixed numbers
    'T11': 99,   # Ratio and proportion
    'T12': 100,  # Percentages and their relationships with fractions and decimals
    'T13': 101,  # Exponential form and calculations using GEMDAS rules
    'T14': 102,  # Common factors, GCF, common multiples, and LCM
    'T15': 103   # Pie graphs
}

topic_formulas = {
    89: "\\sum \\theta_{\\text{vertex}} = 360^\\circ",
    90: "(x, y) \\longrightarrow (x + a, y + b)",
    91: "1 \\text{ L} = 1000 \\text{ cm}^3 = 1000 \\text{ mL}",
    92: "V_{\\text{prism}} = l \\times w \\times h, \\quad V_{\\text{cube}} = s^3",
    93: "A_{\\text{triangle}} = \\frac{1}{2}bh, \\quad A_{\\text{trap}} = \\frac{1}{2}(a+b)h",
    94: "C = 2\\pi r = \\pi d",
    95: "A = \\pi r^2",
    96: "A_{\\text{composite}} = A_1 + A_2",
    97: "a.bcd + e.fgh = \\text{Align Decimal Points}",
    98: "\\frac{a}{b} + \\frac{c}{d} = \\frac{ad + bc}{bd}",
    99: "\\frac{a}{b} = \\frac{c}{d} \\implies a \\cdot d = b \\cdot c",
    100: "\\text{Percentage} = \\frac{\\text{Part}}{\\text{Whole}} \\times 100\\%",
    101: "b^n = \\underbrace{b \\times b \\times \\dots \\times b}_{n \\text{ times}}",
    102: "\\text{GCF}(a, b) = \\text{Greatest Common Factor}",
    103: "\\text{Central Angle} = \\frac{\\text{Category Value}}{\\text{Total Value}} \\times 360^\\circ"
}

def generate_distractors(correct_ans, topic_id, item_num, items_in_topic):
    opts = [correct_ans]
    
    # 1. Check if numerical answer
    clean_ans = correct_ans.replace('°', '').replace('cm²', '').replace('sq cm', '').replace('cm³', '').replace('mL', '').replace('L', '').replace('%', '').replace('m²', '').replace('cm', '').strip()
    
    try:
        val = float(clean_ans)
        if val == int(val):
            val_i = int(val)
            d1 = val_i + 2 if val_i > 0 else 5
            d2 = max(1, val_i - 3) if val_i > 3 else val_i + 4
            d3 = val_i * 2 if val_i > 1 else val_i + 7
            unit = correct_ans[len(clean_ans):]
            cands = [f"{d1}{unit}", f"{d2}{unit}", f"{d3}{unit}"]
            for c in cands:
                if c not in opts:
                    opts.append(c)
        else:
            d1 = round(val + 1.25, 4)
            d2 = round(max(0.0001, val - 0.75), 4)
            d3 = round(val * 1.5, 4)
            unit = correct_ans[len(clean_ans):]
            cands = [f"{d1}{unit}", f"{d2}{unit}", f"{d3}{unit}"]
            for c in cands:
                if c not in opts:
                    opts.append(c)
    except Exception:
        pass
    
    # 2. Use answers from other items in the same topic or domain as distractors
    topic_answers = [it['ans'] for it in items_in_topic if it['ans'] != correct_ans]
    rng = random.Random(item_num + topic_id * 1000)
    rng.shuffle(topic_answers)
    
    for ta in topic_answers:
        if len(opts) >= 4:
            break
        if ta not in opts and len(ta) < 120:
            opts.append(ta)
            
    # 3. Fallback generic topic distractors
    fallback_pool = [
        "No gaps and overlapping allowed at vertices",
        "Perpendicular axis reflection only",
        "Volume measured in square units",
        "Circumference divided by radius squared",
        "Add numerators without common denominator",
        "Part-to-whole ratio only",
        "Multiply base by exponent",
        "Least common multiple instead of GCF",
        "Central angle sum of 180°",
        "Standard metric conversion by factor of 10"
    ]
    for fb in fallback_pool:
        if len(opts) >= 4:
            break
        if fb not in opts:
            opts.append(fb)
            
    while len(opts) < 4:
        opts.append(f"Option {len(opts)+1}")
        
    final_opts = opts[:4]
    rng.shuffle(final_opts)
    return final_opts

# Group items by topic code
items_by_tcode = {}
for item in items:
    tc = item['tcode']
    if tc not in items_by_tcode:
        items_by_tcode[tc] = []
    items_by_tcode[tc].append(item)

generated_questions = []
base_id = 139483  # Match existing DB ID range (139483..140232)

for idx, item in enumerate(items):
    q_id = base_id + idx
    tcode = item['tcode']
    topic_id = topic_map.get(tcode, 89)
    item_id = item['item_id']
    stmt = item['stmt']
    comp = item['comp']
    topic_name = item['topic_name']
    ans = item['ans']
    expl = item['expl']
    
    formula = topic_formulas.get(topic_id, "\\sum \\theta = 360^\\circ")
    
    # Format hint
    hint = f"Recall the concept for {topic_name}: {expl.split('.')[0]}."
    if len(hint) > 160:
        hint = hint[:157] + "..."
        
    # Format working steps from explanation
    expl_sentences = [s.strip() for s in expl.split('.') if s.strip()]
    step1_desc = expl_sentences[0] if len(expl_sentences) > 0 else "Understand problem statement."
    step2_desc = expl_sentences[1] if len(expl_sentences) > 1 else "Apply MATATAG Grade 6 mathematical rules."
    
    steps = [
        f"**Step 1: Analyze problem and key concept**",
        f"{step1_desc}.",
        f"**Step 2: Execute step-by-step mathematical solution**",
        f"{step2_desc}.",
        f"**Final Verified Answer:** {ans}"
    ]
    
    # Generate 4 options with exact correct answer included
    opts = generate_distractors(ans, topic_id, idx, items_by_tcode.get(tcode, []))
    
    short_title = stmt[:45] + "..." if len(stmt) > 45 else stmt
    title = f"[{item_id}] {short_title}"
    
    q_obj = {
        "id": q_id,
        "topic_id": topic_id,
        "question_title": title,
        "question_text": stmt,
        "math_formula": formula,
        "question_type": "MCQ",
        "options_json": json.dumps(opts, ensure_ascii=False),
        "correct_answer": ans,
        "hint": hint,
        "working_steps_json": json.dumps(steps, ensure_ascii=False),
        "image_url": f"/images/g6_t{topic_id}_q{(idx%50)+1}.svg",
        "image_alt": f"Grade 6 {topic_name} Math Diagram",
        "difficulty": ((idx % 3) + 2),
        "created_by": "matatag_g6_verified_excel_generator",
        "created_at": "2026-09-02 22:00:00",
        "updated_at": "2026-09-02 22:00:00"
    }
    generated_questions.append(q_obj)

print(f"Successfully generated {len(generated_questions)} Grade 6 question objects with verified answers!")

# 1. Update questions.json in root and QBank
root_qjson = 'questions.json'
qbank_qjson = os.path.join('QBank', 'questions.json')

for qjson_file in [root_qjson, qbank_qjson]:
    if os.path.exists(qjson_file):
        with open(qjson_file, 'r', encoding='utf-8') as f:
            existing_all = json.load(f)
        
        non_g6 = [q for q in existing_all if not (89 <= q.get('topic_id', 0) <= 103)]
        updated_all = non_g6 + generated_questions
        
        with open(qjson_file, 'w', encoding='utf-8') as f:
            json.dump(updated_all, f, indent=2, ensure_ascii=False)
        print(f"Updated {qjson_file}: Total questions = {len(updated_all)} (Grade 6 = {len(generated_questions)})")

# 2. Update SQLite DBs (server/etuition.db and QBank/server/etuition.db)
server_db = os.path.join('server', 'etuition.db')
qbank_db = os.path.join('QBank', 'server', 'etuition.db')

for db_file in [server_db, qbank_db]:
    if os.path.exists(db_file):
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM questions WHERE topic_id >= 89 AND topic_id <= 103")
        
        for q in generated_questions:
            cursor.execute("""
                INSERT OR REPLACE INTO questions (
                    id, topic_id, question_title, question_text, math_formula,
                    question_type, options_json, correct_answer, hint,
                    working_steps_json, image_url, image_alt, difficulty,
                    created_by, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                q['id'], q['topic_id'], q['question_title'], q['question_text'],
                q['math_formula'], q['question_type'], q['options_json'],
                q['correct_answer'], q['hint'], q['working_steps_json'],
                q['image_url'], q['image_alt'], q['difficulty'],
                q['created_by'], q['created_at'], q['updated_at']
            ))
        
        conn.commit()
        conn.close()
        print(f"Updated SQLite database {db_file} with {len(generated_questions)} Grade 6 questions.")

# 3. Export to Grade_6_Math_50_Unique_Questions_Generated.xlsx and questions_bank.xlsx
out_excel = os.path.join('QBank', 'resources', 'Grade_6_Math_50_Unique_Questions_Generated.xlsx')
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Generated 750 Questions"

headers = [
    "Question ID", "Form Level", "Curriculum Strand", "Unit Title",
    "Topic ID", "Topic Name", "DepEd Competency Code", "Question Title",
    "Question Text", "LaTeX Formula Expression", "Question Type",
    "Option A", "Option B", "Option C", "Option D", "Correct Answer",
    "Hint", "Working Steps", "Image URL", "Image ALT", "Difficulty", "Item ID"
]
ws_out.append(headers)

for q in generated_questions:
    opts = json.loads(q['options_json'])
    opt_a = opts[0] if len(opts) > 0 else ""
    opt_b = opts[1] if len(opts) > 1 else ""
    opt_c = opts[2] if len(opts) > 2 else ""
    opt_d = opts[3] if len(opts) > 3 else ""
    
    t_id = q['topic_id']
    if 89 <= t_id <= 96:
        strand = "Measurement and Geometry"
    elif 97 <= t_id <= 102:
        strand = "Number and Algebra"
    else:
        strand = "Data and Probability"

    row = [
        q['id'], "Form 6", strand, f"Topic {t_id}",
        t_id, f"Topic {t_id}", f"MATATAG-M6-{t_id}", q['question_title'],
        q['question_text'], q['math_formula'], q['question_type'],
        opt_a, opt_b, opt_c, opt_d, q['correct_answer'],
        q['hint'], q['working_steps_json'], q['image_url'], q['image_alt'],
        q['difficulty'], q['question_title'].split(']')[0].replace('[', '')
    ]
    ws_out.append(row)

wb_out.save(out_excel)
print(f"Saved generated Excel workbook to {out_excel}")

# Export full questions bank
out_qb_excel = os.path.join('QBank', 'questions_bank.xlsx')
if os.path.exists(server_db):
    conn = sqlite3.connect(server_db)
    c = conn.cursor()
    c.execute('''
        SELECT q.id, t.form_level, t.strand, t.unit, q.topic_id, t.title,
               t.competencies, q.question_title, q.question_text, q.math_formula,
               q.question_type, q.options_json, q.correct_answer, q.hint,
               q.working_steps_json, q.image_url, q.image_alt, q.difficulty
        FROM questions q
        JOIN topics t ON q.topic_id = t.id
        ORDER BY q.topic_id, q.id
    ''')
    all_db_rows = c.fetchall()
    conn.close()

    wb_qb = openpyxl.Workbook()
    ws_qb = wb_qb.active
    ws_qb.title = 'Question Bank'

    qb_headers = [
        'Question ID', 'Form Level', 'Curriculum Strand', 'Unit Title',
        'Topic ID', 'Topic Name', 'DepEd Competency Code', 'Question Title',
        'Question Text', 'LaTeX Formula Expression', 'Question Type',
        'Option A', 'Option B', 'Option C', 'Option D', 'Correct Answer',
        'Hint', 'Working Steps', 'Image URL', 'Image ALT', 'Difficulty'
    ]
    ws_qb.append(qb_headers)

    for r in all_db_rows:
        opts = json.loads(r[11]) if r[11] else []
        opt_a = opts[0] if len(opts) > 0 else ''
        opt_b = opts[1] if len(opts) > 1 else ''
        opt_c = opts[2] if len(opts) > 2 else ''
        opt_d = opts[3] if len(opts) > 3 else ''
        
        ws_qb.append([
            r[0], f'Form {r[1]}', r[2], r[3], r[4], r[5], r[6], r[7],
            r[8], r[9], r[10], opt_a, opt_b, opt_c, opt_d, r[12],
            r[13], r[14], r[15], r[16], r[17]
        ])

    wb_qb.save(out_qb_excel)
    print(f"Exported {len(all_db_rows)} total questions to {out_qb_excel}")

print("✨ Grade 6 Question Bank update with verified Excel answers completed successfully!")
