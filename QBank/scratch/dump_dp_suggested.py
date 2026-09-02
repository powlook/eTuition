import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('resources/Grade_10_Math_Questions.xlsx', read_only=True)
sheet_name = [s for s in wb.sheetnames if 'Data' in s and 'Probability' in s][0]
sheet = wb[sheet_name]

topics = {}
for r in sheet.iter_rows(min_row=5, values_only=True):
    qid = r[0]
    topic_str = r[1]
    domain = r[2]
    comp = r[3]
    qtext = r[4]
    if topic_str and qtext:
        tcode = topic_str.split(':')[0].strip()
        if tcode not in topics:
            topics[tcode] = {
                'topic_code': tcode,
                'topic_name': topic_str.split(':', 1)[1].strip() if ':' in topic_str else topic_str,
                'domain': domain,
                'competency': comp,
                'questions': []
            }
        topics[tcode]['questions'].append({
            'item_id': qid,
            'suggested_question': qtext
        })

with open('scratch/g10_dp_suggested.json', 'w', encoding='utf-8') as f:
    json.dump(topics, f, indent=2, ensure_ascii=False)

with open('scratch/all_dp_suggested.txt', 'w', encoding='utf-8') as f:
    for tcode, tdata in topics.items():
        f.write(f"\n==========================================\n")
        f.write(f"TOPIC {tcode}: {tdata['topic_name']}\n")
        f.write(f"==========================================\n")
        for q in tdata['questions']:
            f.write(f"[{q['item_id']}] {q['suggested_question']}\n")

print(f"Extracted {len(topics)} topics with question counts:")
for tcode, data in topics.items():
    print(f"  {tcode} ({data['topic_name']}): {len(data['questions'])} questions")
