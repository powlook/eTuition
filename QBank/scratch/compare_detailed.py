import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb_qb = openpyxl.load_workbook('questions_bank.xlsx', read_only=True)
sheet_qb = wb_qb['DepEd Question Bank']

qb_by_topic = {}
for row in sheet_qb.iter_rows(min_row=2, values_only=True):
    tid = row[4]
    if tid in [161, 162, 163, 164]:
        if tid not in qb_by_topic:
            qb_by_topic[tid] = []
        qb_by_topic[tid].append(row)

for tid, qlist in qb_by_topic.items():
    print(f"\nquestions_bank.xlsx Topic ID {tid}: count = {len(qlist)}")
    print("First 3 questions in questions_bank.xlsx:")
    for r in qlist[:3]:
        print(f"  QID: {r[0]}, Title: {r[7]}, Text: {r[8][:70]}...")

wb_g10 = openpyxl.load_workbook('resources/Grade_10_Math_Questions.xlsx', read_only=True)
sheet_g10 = wb_g10['Measurement & Geometry (MG)']
g10_by_topic = {}
for row in sheet_g10.iter_rows(min_row=5, values_only=True):
    if row[1]:
        tcode = row[1].split(':')[0]
        if tcode not in g10_by_topic:
            g10_by_topic[tcode] = []
        g10_by_topic[tcode].append(row)

for tcode, qlist in g10_by_topic.items():
    print(f"\nGrade_10_Math_Questions.xlsx {tcode}: count = {len(qlist)}")
    print("First 3 questions in Grade_10_Math_Questions.xlsx:")
    for r in qlist[:3]:
        print(f"  ItemID: {r[0]}, Question: {r[4][:70]}...")
