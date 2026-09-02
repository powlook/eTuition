import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('resources/Grade_10_Math_Questions.xlsx')
sheet = wb['Measurement & Geometry (MG)']

topics = {}
for r in range(5, sheet.max_row + 1):
    qid = sheet.cell(r, 1).value
    topic = sheet.cell(r, 2).value
    domain = sheet.cell(r, 3).value
    comp = sheet.cell(r, 4).value
    qtext = sheet.cell(r, 5).value
    if topic:
        if topic not in topics:
            topics[topic] = {'comp': comp, 'questions': []}
        topics[topic]['questions'].append((qid, qtext))

print(f"Found {len(topics)} topics in Measurement & Geometry (MG):")
for t, d in topics.items():
    print(f"\nTopic: {t}")
    print(f"Competency: {d['comp']}")
    print(f"Total questions present: {len(d['questions'])}")
    print("First 3 questions:")
    for qid, qt in d['questions'][:3]:
        print(f"  [{qid}] {qt}")
    print("Last 3 questions:")
    for qid, qt in d['questions'][-3:]:
        print(f"  [{qid}] {qt}")
