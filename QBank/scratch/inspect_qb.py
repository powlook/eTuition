import openpyxl

qb_path = 'questions_bank.xlsx'
wb = openpyxl.load_workbook(qb_path, read_only=True)
print("Sheet names in questions_bank.xlsx:", wb.sheetnames)

sheet = wb.active
print("Active sheet title:", sheet.title)

# Read header row
headers = []
for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    headers = list(row)
print("\nHeaders in questions_bank.xlsx:")
for idx, h in enumerate(headers):
    print(f"  Col {idx+1}: {h}")

count = 0
g10_count = 0
mg_g10_count = 0

for row in sheet.iter_rows(min_row=2, values_only=True):
    count += 1
    form_level = row[1] if len(row) > 1 else None
    strand = row[2] if len(row) > 2 else None
    if str(form_level) in ['10', 'Form 10']:
        g10_count += 1
        if strand and 'Measurement' in str(strand):
            mg_g10_count += 1

print(f"\nTotal rows in questions_bank.xlsx: {count}")
print(f"Form 10 rows: {g10_count}")
print(f"Form 10 Measurement & Geometry rows: {mg_g10_count}")
