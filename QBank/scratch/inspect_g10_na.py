import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('resources/Grade_10_Math_Questions.xlsx', read_only=True)
sheet_name = [s for s in wb.sheetnames if 'Number' in s and 'Algebra' in s][0]
print(f"Sheet found: '{sheet_name}'")

sheet = wb[sheet_name]

topics = {}
for r in sheet.iter_rows(min_row=5, values_only=True):
    qid = r[0]
    topic_str = r[1]
    domain = r[2]
    comp = r[3]
    qtext = r[4]
    if topic_str and qtext:
        tcode = topic_str.split(':')[0].strip()
        if tcode not in topics:
            topics[tcode] = {
                'topic_code': tcode,
                'topic_name': topic_str.split(':', 1)[1].strip() if ':' in topic_str else topic_str,
                'domain': domain,
                'competency': comp,
                'questions': []
            }
        topics[tcode]['questions'].append({
            'item_id': qid,
            'suggested_question': qtext
        })

print(f"Found {len(topics)} topics in {sheet_name}:")
for tcode, data in topics.items():
    print(f"  {tcode} ({data['topic_name']}): {len(data['questions'])} questions")

with open('scratch/g10_na_suggested.json', 'w', encoding='utf-8') as f:
    json.dump(topics, f, indent=2, ensure_ascii=False)
