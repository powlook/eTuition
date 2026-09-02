import openpyxl
import sqlite3
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

db_path = 'server/qbank.db'
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Retrieve questions for topics 161, 162, 163, 164
cursor.execute("""
    SELECT q.id, t.id, t.title, t.competencies, q.question_title, q.question_text, q.options_json, q.correct_answer, q.working_steps_json, q.image_url
    FROM questions q
    JOIN topics t ON q.topic_id = t.id
    WHERE t.id IN (161, 162, 163, 164)
    ORDER BY t.id ASC, q.id ASC
""")
rows = cursor.fetchall()
print(f"Fetched {len(rows)} questions from DB for Form 10 MG (Topics 161..164).")

topic_code_map = {
    161: "T01: The laws of sines and the laws of cosines",
    162: "T02: Translations, reflections, and rotations in the Cartesian plane",
    163: "T03: Central angles, inscribed angles, chords, secants, and tangents",
    164: "T04: Sectors and segments of a circle, and their areas"
}

excel_path = 'resources/Grade_10_Math_Questions.xlsx'
wb = openpyxl.load_workbook(excel_path)

# Update Sheet 'Measurement & Geometry (MG)'
mg_sheet_name = 'Measurement & Geometry (MG)' if 'Measurement & Geometry (MG)' in wb.sheetnames else 'Measurement and Geometry (MG)'
sheet_mg = wb[mg_sheet_name]

# Map topic id to question count for item_id generation
t_count = {161: 0, 162: 0, 163: 0, 164: 0}

for row_idx, r in enumerate(rows, start=5):
    qid, tid, ttitle, comp, qtitle, qtext, opt_json, ans, steps_json, img_url = r
    t_count[tid] += 1
    t_code_short = f"T{tid - 160:02d}"
    item_id = f"{t_code_short}-Q{t_count[tid]:03d}"
    
    t_code_name = topic_code_map[tid]
    domain = "Measurement and Geometry (MG)"
    
    sheet_mg.cell(row=row_idx, column=1, value=item_id)
    sheet_mg.cell(row=row_idx, column=2, value=t_code_name)
    sheet_mg.cell(row=row_idx, column=3, value=domain)
    sheet_mg.cell(row=row_idx, column=4, value=comp)
    sheet_mg.cell(row=row_idx, column=5, value=qtext)

# Also update 'All 800 Questions' sheet if present
if 'All 800 Questions' in wb.sheetnames:
    sheet_all = wb['All 800 Questions']
    # Find matching Item ID or update starting rows for T01-Q001 to T04-Q050
    item_row_map = {}
    for r_idx in range(5, sheet_all.max_row + 1):
        val = sheet_all.cell(row=r_idx, column=1).value
        if val:
            item_row_map[str(val).strip()] = r_idx
    
    t_count = {161: 0, 162: 0, 163: 0, 164: 0}
    for r in rows:
        qid, tid, ttitle, comp, qtitle, qtext, opt_json, ans, steps_json, img_url = r
        t_count[tid] += 1
        t_code_short = f"T{tid - 160:02d}"
        item_id = f"{t_code_short}-Q{t_count[tid]:03d}"
        
        target_r = item_row_map.get(item_id)
        if target_r:
            sheet_all.cell(row=target_r, column=5, value=qtext)

wb.save(excel_path)
print(f"✅ Successfully updated {excel_path} sheet '{mg_sheet_name}' and 'All 800 Questions' with complete questions!")
