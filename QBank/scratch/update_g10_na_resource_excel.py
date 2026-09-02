import openpyxl
import sqlite3
import json
import time
import sys

sys.stdout.reconfigure(encoding='utf-8')

db_path = 'server/qbank.db'
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Retrieve questions for topics 165..172
cursor.execute("""
    SELECT q.id, t.id, t.title, t.competencies, q.question_title, q.question_text, q.options_json, q.correct_answer, q.working_steps_json, q.image_url
    FROM questions q
    JOIN topics t ON q.topic_id = t.id
    WHERE t.id BETWEEN 165 AND 172
    ORDER BY t.id ASC, q.id ASC
""")
rows = cursor.fetchall()
conn.close()
print(f"Fetched {len(rows)} questions from DB for Form 10 NA (Topics 165..172).")

topic_code_map = {
    165: "T05: Quadratic inequalities in one variable and in two variables",
    166: "T06: Absolute value equations and inequalities in one variable and their graphs",
    167: "T07: Radical expressions",
    168: "T08: The roots of a quadratic equation",
    169: "T09: Quadratic functions",
    170: "T10: Equations reducible to quadratic equations",
    171: "T11: Equation of a circle and the graph of a circle",
    172: "T12: Simple interest, compound interest, and depreciation"
}

excel_path = 'resources/Grade_10_Math_Questions.xlsx'
wb = openpyxl.load_workbook(excel_path)

# Update Sheet 'Number and Algebra (NA)'
na_sheet_name = 'Number and Algebra (NA)' if 'Number and Algebra (NA)' in wb.sheetnames else 'Number & Algebra (NA)'
sheet_na = wb[na_sheet_name]

# Map topic id to question count for item_id generation
t_count = {tid: 0 for tid in range(165, 173)}

for row_idx, r in enumerate(rows, start=5):
    qid, tid, ttitle, comp, qtitle, qtext, opt_json, ans, steps_json, img_url = r
    t_count[tid] += 1
    t_code_short = f"T{tid - 160:02d}"
    item_id = f"{t_code_short}-Q{t_count[tid]:03d}"
    
    t_code_name = topic_code_map[tid]
    domain = "Number and Algebra (NA)"
    
    sheet_na.cell(row=row_idx, column=1, value=item_id)
    sheet_na.cell(row=row_idx, column=2, value=t_code_name)
    sheet_na.cell(row=row_idx, column=3, value=domain)
    sheet_na.cell(row=row_idx, column=4, value=comp)
    sheet_na.cell(row=row_idx, column=5, value=qtext)

# Also update 'All 800 Questions' sheet if present
if 'All 800 Questions' in wb.sheetnames:
    sheet_all = wb['All 800 Questions']
    # Find matching Item ID or update starting rows for T05-Q001 to T12-Q050
    item_row_map = {}
    for r_idx in range(5, sheet_all.max_row + 1):
        val = sheet_all.cell(row=r_idx, column=1).value
        if val:
            item_row_map[str(val).strip()] = r_idx
    
    t_count = {tid: 0 for tid in range(165, 173)}
    for r in rows:
        qid, tid, ttitle, comp, qtitle, qtext, opt_json, ans, steps_json, img_url = r
        t_count[tid] += 1
        t_code_short = f"T{tid - 160:02d}"
        item_id = f"{t_code_short}-Q{t_count[tid]:03d}"
        
        target_r = item_row_map.get(item_id)
        if target_r:
            sheet_all.cell(row=target_r, column=5, value=qtext)

wb.save(excel_path)
wb.close()
print(f"✅ Successfully updated {excel_path} sheet '{na_sheet_name}' and 'All 800 Questions' with complete questions!")
