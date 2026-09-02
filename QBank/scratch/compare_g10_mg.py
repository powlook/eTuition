import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Inspect questions_bank.xlsx for Form 10 MG
wb_qb = openpyxl.load_workbook('questions_bank.xlsx', read_only=True)
sheet_qb = wb_qb['DepEd Question Bank']

qb_rows = []
for row in sheet_qb.iter_rows(min_row=2, values_only=True):
    form_level = str(row[1]) if row[1] is not None else ''
    strand = str(row[2]) if row[2] is not None else ''
    if form_level in ['10', 'Form 10'] and 'Measurement' in strand:
        qb_rows.append(row)

print(f"questions_bank.xlsx Form 10 MG count: {len(qb_rows)}")
if qb_rows:
    print("Sample question from questions_bank.xlsx:")
    r = qb_rows[0]
    print(f"  ID: {r[0]}, Topic ID: {r[4]}, Topic Name: {r[5]}")
    print(f"  Title: {r[7]}")
    print(f"  Text: {r[8]}")
    print(f"  Opt A: {r[11]}, Opt B: {r[12]}, Opt C: {r[13]}, Opt D: {r[14]}")
    print(f"  Answer: {r[15]}")
    print(f"  Working: {r[17][:150]}...")
    print(f"  Image: {r[18]}")

# Inspect Grade_10_Math_Questions.xlsx for MG sheet
wb_g10 = openpyxl.load_workbook('resources/Grade_10_Math_Questions.xlsx', read_only=True)
sheet_g10 = wb_g10['Measurement & Geometry (MG)']

g10_rows = []
for row in sheet_g10.iter_rows(min_row=5, values_only=True):
    if any(row):
        g10_rows.append(row)

print(f"\nGrade_10_Math_Questions.xlsx MG sheet count: {len(g10_rows)}")
if g10_rows:
    print("Sample row from Grade_10_Math_Questions.xlsx:")
    r = g10_rows[0]
    print(f"  Item ID: {r[0]}")
    print(f"  Topic Code & Name: {r[1]}")
    print(f"  Domain: {r[2]}")
    print(f"  Competency: {r[3]}")
    print(f"  Question: {r[4]}")
